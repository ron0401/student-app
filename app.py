import streamlit as st
import pandas as pd
import openpyxl
import itertools
from io import BytesIO
from tempfile import NamedTemporaryFile


st.title('成績整理ちゃん')

uploaded_file = st.file_uploader("ファイルをアップロードしてください", type='xlsx')

if uploaded_file is not None:
    excel_book = uploaded_file.name
    sheet1 = "表紙"
    sheet2 = "入力"

    cells = "B4:B8"

    book = openpyxl.load_workbook(excel_book)
    sheet = book[sheet1]
    kyoka_list = []
    for row in sheet.iter_rows(min_row=4, max_row=20, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:
                kyoka_list.append(cell.value)

    shubetu_list = []
    for row in sheet.iter_rows(min_row=4, max_row=20, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                shubetu_list.append(cell.value)

    students = []
    for row in sheet.iter_rows(min_row=9, max_row=100, min_col=6, max_col=6):
        for cell in row:
            if cell.value is not None:
                students.append(cell.value)
    
    kyoka_shubetu_list = list(itertools.product(kyoka_list, shubetu_list))

    class unit_score:
        def __init__(self, student, score):
            self.student = student
            self.score = score


    class unit_test:
        def __init__(self, kyoka, shubetu, scores: list[unit_score], weight, max_score, name):
            self.kyoka = kyoka
            self.shubetu = shubetu
            self.scores = scores
            self.weight = weight
            self.max_score = max_score
            self.name = name
        
        @property
        def score_result(self) -> dict:
            result = {}
            for score in self.scores:
                result[score.student] = score.score * self.weight / self.max_score
            return result
        
        def __repr__(self) -> str:
            return f"Test {self.name} for {self.kyoka} {self.shubetu} with {len(self.scores)} students"


    tests: list[unit_test] = []
    for i in book[sheet2].iter_cols(min_row=1, max_row=100, min_col=3, max_col=153):
        lst = []
        for j in i:
            lst.append(j.value)
        if lst[1] is not None:
            scores = []
            for k in range(7, 7 + len(students)):
                scores.append(unit_score(students[k - 7], lst[k]))
            tests.append(unit_test(lst[2], lst[3], scores, lst[4], lst[5], lst[1]))

    class unit_seiseki:
        def __init__(self, kyoka, shubetu, tests: list[unit_test], students: list[str]):
            self.kyoka = kyoka
            self.shubetu = shubetu
            self.tests = tests
            self.students = students

        @property
        def data(self):
            df = pd.DataFrame()
            df["student"] = self.students
            total_score = dict.fromkeys(self.students, 0)
            for i in total_score.keys():
                total_score[i] = 0
            for test in self.tests:
                result = test.score_result
                for r in result.keys():
                    total_score[r] += result[r]
            
            df["total_score"] = total_score.values()
            df.set_index("student", inplace=True)
            df["ranking"] = [int(i) for i in df["total_score"].rank(ascending=False)]
            df = df.sort_values(by="total_score", ascending=False)
            rank_list = ["A", "B", "C"]
            ## 30% of students are A, 40% of students are B, 30% of students are C
            rank = []
            for i in range(len(df)):
                if i < len(df) * 0.3:
                    rank.append("A")
                elif i < len(df) * 0.7:
                    rank.append("B")
                else:
                    rank.append("C")
            df["rank"] = rank
            df.reset_index(inplace=True)
            # print(len(self.tests[0].score_result))
            return df



    seiseki_list: list[unit_seiseki] = []

    for kyoka, shubetu in kyoka_shubetu_list:
        tests_for_seiseki = []
        for test in tests:
            if test.kyoka == kyoka and test.shubetu == shubetu:
                tests_for_seiseki.append(test)
        seiseki_list.append(unit_seiseki(kyoka, shubetu, tests_for_seiseki, students))

    result_book = openpyxl.Workbook()
    result_book.create_sheet(title="result")
    result_book.remove(result_book["Sheet"])
    result_sheet = result_book["result"]
    start_col = 3
    start_row = 5

    for i in seiseki_list:
        df = i.data.copy()
        result_sheet.cell(row=2, column=start_col, value="教科：")
        result_sheet.cell(row=2, column=start_col+1, value=i.kyoka)
        result_sheet.cell(row=3, column=start_col, value="種別：")
        result_sheet.cell(row=3, column=start_col+1, value=i.shubetu)
        result_sheet.cell(row=start_row-1, column=start_col, value="生徒")
        result_sheet.cell(row=start_row-1, column=start_col+1, value="合計点")
        result_sheet.cell(row=start_row-1, column=start_col+2, value="順位")
        result_sheet.cell(row=start_row-1, column=start_col+3, value="ランク")
        for r in df.values:
            result_sheet.cell(row=start_row, column=start_col, value=r[0])
            result_sheet.cell(row=start_row, column=start_col + 1, value=r[1])
            result_sheet.cell(row=start_row, column=start_col + 2, value=r[2])
            result_sheet.cell(row=start_row, column=start_col + 3, value=r[3])

            start_row += 1
        
        start_row = 5
        start_col += 5
        
    
    kyoka_selected = st.selectbox("教科", kyoka_list)
    shubetu_selected = st.selectbox("種別", shubetu_list)

    filename = "成績表.xlsx"
    with NamedTemporaryFile() as tmp:
        result_book.save(tmp.name)
        data = tmp.read()

    
    if kyoka_selected is not None and shubetu_selected is not None:
        for i in seiseki_list:
            if i.kyoka == kyoka_selected and i.shubetu == shubetu_selected:
                st.write(i.data)
                break



    st.download_button(
        label="成績表ダウンロード",
        data=BytesIO(data),
        file_name="成績表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


